"""
A题 Q3~Q5 求解 + result文件输出
"""
import numpy as np
import openpyxl
from core import *
from optimizer import calc_time_fast, calc_time_precise, calc_multi_bomb_time, solve_q2

# ============ Q3: FY1投3弹,分层优化 ============

def solve_q3(q2_params):
    """Q3: FY1投3弹对M1干扰
    策略: 第1弹沿用Q2最优, 优化第2/3弹时序
    """
    theta, v = q2_params[0], q2_params[1]
    t_rel1, dt_fuse1 = q2_params[2], q2_params[3]

    print("Q3: FY1三弹分层优化")
    print(f"  第1弹(沿用Q2): θ={np.degrees(theta):.2f}°, v={v:.1f}, t_rel={t_rel1:.3f}, dt={dt_fuse1:.3f}")

    rng = np.random.RandomState(123)
    n_particles = 25
    n_iter = 35

    # 变量: t_rel2, dt_fuse2, t_rel3, dt_fuse3
    lb = np.array([t_rel1 + 1.0, 0.0, t_rel1 + 2.0, 0.0])
    ub = np.array([15.0, 6.0, 20.0, 6.0])
    dim = 4

    pos = rng.uniform(lb, ub, (n_particles, dim))
    vel = np.zeros((n_particles, dim))
    pbest = pos.copy()

    def fitness(x):
        tr2, df2, tr3, df3 = x
        # 检查约束
        if tr2 < t_rel1 + 1.0 or tr3 < tr2 + 1.0:
            return 0.0
        if UAVS['FY1'][2] - 0.5 * G * df2 ** 2 < 0:
            return 0.0
        if UAVS['FY1'][2] - 0.5 * G * df3 ** 2 < 0:
            return 0.0
        bombs = [(t_rel1, dt_fuse1), (tr2, df2), (tr3, df3)]
        T, _, _ = calc_multi_bomb_time('M1', 'FY1', theta, v, bombs,
                                        dt_step=0.01, n_samples=120)
        return T

    pbest_fit = np.array([fitness(p) for p in pos])
    gbest_idx = np.argmax(pbest_fit)
    gbest = pbest[gbest_idx].copy()
    gbest_fit = pbest_fit[gbest_idx]

    w, c1, c2 = 0.7, 1.5, 1.5
    for it in range(n_iter):
        r1, r2 = rng.random((n_particles, dim)), rng.random((n_particles, dim))
        vel = w * vel + c1 * r1 * (pbest - pos) + c2 * r2 * (gbest - pos)
        pos = np.clip(pos + vel, lb, ub)
        for i in range(n_particles):
            f = fitness(pos[i])
            if f > pbest_fit[i]:
                pbest_fit[i], pbest[i] = f, pos[i].copy()
            if f > gbest_fit:
                gbest_fit, gbest = f, pos[i].copy()
        if (it + 1) % 10 == 0:
            print(f"  Q3 PSO iter {it+1}: gbest={gbest_fit:.4f}s")

    tr2, df2, tr3, df3 = gbest
    bombs = [(t_rel1, dt_fuse1), (tr2, df2), (tr3, df3)]
    T, singles, merged = calc_multi_bomb_time('M1', 'FY1', theta, v, bombs,
                                               dt_step=0.001, n_samples=360)
    print(f"\nQ3结果(精确):")
    for i, (tr, df) in enumerate(bombs):
        print(f"  弹{i+1}: t_rel={tr:.4f}s, dt_fuse={df:.4f}s, 单独={singles[i]:.4f}s")
    print(f"  并集遮蔽时长: {T:.4f}s")
    print(f"  合并区间数: {len(merged)}")
    return theta, v, bombs, T, merged


# ============ Q4: 3机各1弹 ============

def solve_q4(q2_params):
    """Q4: FY1/FY2/FY3各1弹对M1
    策略: 各机独立单弹最优 + 联合微调
    """
    print("Q4: 三机协同(FY1/FY2/FY3各1弹)")

    # FY1用Q2最优
    theta1, v1, tr1, df1 = q2_params

    # FY2, FY3各自独立优化
    results = {'FY1': (theta1, v1, tr1, df1)}

    for uav_id in ['FY2', 'FY3']:
        print(f"  优化 {uav_id}...")
        rng = np.random.RandomState(hash(uav_id) % 1000)
        n_p, n_i = 20, 30
        lb = np.array([0.0, 70.0, 0.0, 0.0])
        ub = np.array([2 * np.pi, 140.0, 10.0, 6.0])
        pos = rng.uniform(lb, ub, (n_p, 4))
        vel = np.zeros((n_p, 4))
        pbest = pos.copy()

        def fit(x, uid=uav_id):
            th, vv, tr, df = x
            if UAVS[uid][2] - 0.5 * G * df ** 2 < 0:
                return 0.0
            T, _, _, _ = calc_time_fast('M1', uid, th, vv, tr, df,
                                         dt_step=0.01, n_samples=120)
            return T

        pf = np.array([fit(p) for p in pos])
        gi = np.argmax(pf)
        gb = pos[gi].copy()
        gf = pf[gi]

        for it in range(n_i):
            r1, r2 = rng.random((n_p, 4)), rng.random((n_p, 4))
            vel = 0.7 * vel + 1.5 * r1 * (pbest - pos) + 1.5 * r2 * (gb - pos)
            pos = np.clip(pos + vel, lb, ub)
            for i in range(n_p):
                f = fit(pos[i])
                if f > pf[i]:
                    pf[i], pbest[i] = f, pos[i].copy()
                if f > gf:
                    gf, gb = f, pos[i].copy()

        results[uav_id] = tuple(gb)
        print(f"    {uav_id}: θ={np.degrees(gb[0]):.2f}°, v={gb[1]:.1f}, T={gf:.3f}s")

    # 联合并集计算
    all_intervals = []
    for uav_id, (th, vv, tr, df) in results.items():
        T, intervals, _, _ = calc_time_precise('M1', uav_id, th, vv, tr, df)
        all_intervals.extend(intervals)
        print(f"  {uav_id}: T_single={T:.4f}s, 区间数={len(intervals)}")

    all_intervals.sort()
    merged = [list(all_intervals[0])]
    for s, e in all_intervals[1:]:
        if s <= merged[-1][1]:
            merged[-1][1] = max(merged[-1][1], e)
        else:
            merged.append([s, e])
    union_T = sum(e - s for s, e in merged)
    print(f"\nQ4结果: 联合遮蔽时长={union_T:.4f}s, 合并区间数={len(merged)}")
    return results, union_T, merged


# ============ Q5: 5机至多3弹→M1M2M3 ============

def solve_q5(q2_params, q4_results):
    """Q5: 5机各至多3弹,对M1/M2/M3
    策略: 任务分配(FY1→M1, FY2→M1, FY3→M2, FY4→M2/M3, FY5→M3)
    分组并行优化各机策略
    """
    print("Q5: 5机多对多协同")

    # 任务分配: 每架机负责1枚导弹,各投1~2弹
    # FY1→M1(2弹), FY2→M1(1弹), FY3→M2(2弹), FY4→M3(2弹), FY5→M3(1弹)
    assignment = {
        'FY1': ('M1', 2),
        'FY2': ('M1', 1),
        'FY3': ('M2', 2),
        'FY4': ('M3', 2),
        'FY5': ('M3', 1),
    }

    all_results = {}
    total_T = 0.0

    for uav_id, (missile_id, n_bombs) in assignment.items():
        print(f"  优化 {uav_id}→{missile_id} ({n_bombs}弹)...")

        # 用Q4结果作为FY1/FY2/FY3的初值
        if uav_id in q4_results:
            theta0, v0 = q4_results[uav_id][0], q4_results[uav_id][1]
        else:
            theta0, v0 = np.pi, 100.0

        rng = np.random.RandomState(hash(uav_id + missile_id) % 1000)
        n_p, n_i = 20, 25

        if n_bombs == 1:
            # 单弹优化
            lb = np.array([0.0, 70.0, 0.0, 0.0])
            ub = np.array([2 * np.pi, 140.0, 10.0, 6.0])
            dim = 4
            pos = rng.uniform(lb, ub, (n_p, dim))
            pos[0] = [theta0, v0, 1.0, 2.0]
            vel = np.zeros((n_p, dim))
            pbest = pos.copy()

            def fit(x, uid=uav_id, mid=missile_id):
                th, vv, tr, df = x
                if UAVS[uid][2] - 0.5 * G * df ** 2 < 0:
                    return 0.0
                T, _, _, _ = calc_time_fast(mid, uid, th, vv, tr, df,
                                             dt_step=0.01, n_samples=100)
                return T

            pf = np.array([fit(p) for p in pos])
            gi = np.argmax(pf)
            gb, gf = pos[gi].copy(), pf[gi]

            for it in range(n_i):
                r1, r2 = rng.random((n_p, dim)), rng.random((n_p, dim))
                vel = 0.7 * vel + 1.5 * r1 * (pbest - pos) + 1.5 * r2 * (gb - pos)
                pos = np.clip(pos + vel, lb, ub)
                for i in range(n_p):
                    f = fit(pos[i])
                    if f > pf[i]:
                        pf[i], pbest[i] = f, pos[i].copy()
                    if f > gf:
                        gf, gb = f, pos[i].copy()

            th, vv, tr, df = gb
            T, intervals, bp, tb = calc_time_precise(missile_id, uav_id, th, vv, tr, df)
            all_results[uav_id] = (missile_id, [(th, vv, [(tr, df)], T, intervals)])
            total_T += T
            print(f"    {uav_id}: T={T:.3f}s")

        else:
            # 双弹优化: theta, v共享; 2组(t_rel, dt_fuse)
            lb = np.array([0.0, 70.0, 0.0, 0.0, 1.0, 0.0])
            ub = np.array([2 * np.pi, 140.0, 10.0, 6.0, 20.0, 6.0])
            dim = 6
            pos = rng.uniform(lb, ub, (n_p, dim))
            pos[0] = [theta0, v0, 1.0, 2.0, 3.0, 2.0]
            vel = np.zeros((n_p, dim))
            pbest = pos.copy()

            def fit(x, uid=uav_id, mid=missile_id):
                th, vv, tr1, df1, tr2, df2 = x
                if tr2 < tr1 + 1.0:
                    return 0.0
                if UAVS[uid][2] - 0.5 * G * df1 ** 2 < 0:
                    return 0.0
                if UAVS[uid][2] - 0.5 * G * df2 ** 2 < 0:
                    return 0.0
                bombs = [(tr1, df1), (tr2, df2)]
                T, _, _ = calc_multi_bomb_time(mid, uid, th, vv, bombs,
                                                dt_step=0.01, n_samples=100)
                return T

            pf = np.array([fit(p) for p in pos])
            gi = np.argmax(pf)
            gb, gf = pos[gi].copy(), pf[gi]

            for it in range(n_i):
                r1, r2 = rng.random((n_p, dim)), rng.random((n_p, dim))
                vel = 0.7 * vel + 1.5 * r1 * (pbest - pos) + 1.5 * r2 * (gb - pos)
                pos = np.clip(pos + vel, lb, ub)
                for i in range(n_p):
                    f = fit(pos[i])
                    if f > pf[i]:
                        pf[i], pbest[i] = f, pos[i].copy()
                    if f > gf:
                        gf, gb = f, pos[i].copy()

            th, vv, tr1, df1, tr2, df2 = gb
            bombs = [(tr1, df1), (tr2, df2)]
            T, singles, merged = calc_multi_bomb_time(missile_id, uav_id, th, vv, bombs,
                                                       dt_step=0.005, n_samples=180)
            all_results[uav_id] = (missile_id, [(th, vv, bombs, T, merged)])
            total_T += T
            print(f"    {uav_id}: T={T:.3f}s (弹1={singles[0]:.3f}, 弹2={singles[1]:.3f})")

    print(f"\nQ5结果: 总遮蔽时长={total_T:.4f}s (各导弹独立求和)")
    return all_results, total_T


# ============ 输出result文件 ============

def write_result1(theta, v, bombs, T):
    """result1.xlsx: FY1投3弹对M1"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ['无人机运动方向', '无人机运动速度 (m/s)', '烟幕干扰弹编号',
               '烟幕干扰弹投放点的x坐标 (m)', '烟幕干扰弹投放点的y坐标 (m)', '烟幕干扰弹投放点的z坐标 (m)',
               '烟幕干扰弹起爆点的x坐标 (m)', '烟幕干扰弹起爆点的y坐标 (m)', '烟幕干扰弹起爆点的z坐标 (m)',
               '有效干扰时长 (s)']
    ws.append(headers)
    angle_deg = np.degrees(theta) % 360
    for j, (tr, df) in enumerate(bombs):
        drop = uav_pos('FY1', theta, v, tr)
        bp, _ = burst_point('FY1', theta, v, tr, df)
        T_single, _, _, _ = calc_time_precise('M1', 'FY1', theta, v, tr, df)
        ws.append([angle_deg, v, j + 1, drop[0], drop[1], drop[2],
                   bp[0], bp[1], bp[2], round(T_single, 4)])
    ws.append([None] * 10)
    ws.append(['注：以x轴为正向，逆时针方向为正，取值0~360（度）。'] + [None] * 9)
    path = "C:\\Users\\wuyan\\Desktop\\xiaosai\\A题_solution\\stage2\\result1.xlsx"
    wb.save(path)
    print(f"  result1.xlsx 已保存: {path}")


def write_result2(results):
    """result2.xlsx: 3机各1弹对M1"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ['无人机编号', '无人机运动方向', '无人机运动速度 (m/s)',
               '烟幕干扰弹投放点的x坐标 (m)', '烟幕干扰弹投放点的y坐标 (m)', '烟幕干扰弹投放点的z坐标 (m)',
               '烟幕干扰弹起爆点的x坐标 (m)', '烟幕干扰弹起爆点的y坐标 (m)', '烟幕干扰弹起爆点的z坐标 (m)',
               '有效干扰时长 (s)']
    ws.append(headers)
    for uav_id in ['FY1', 'FY2', 'FY3']:
        th, vv, tr, df = results[uav_id]
        drop = uav_pos(uav_id, th, vv, tr)
        bp, _ = burst_point(uav_id, th, vv, tr, df)
        T, _, _, _ = calc_time_precise('M1', uav_id, th, vv, tr, df)
        ws.append([uav_id, np.degrees(th) % 360, vv, drop[0], drop[1], drop[2],
                   bp[0], bp[1], bp[2], round(T, 4)])
    ws.append([None] * 10)
    ws.append([None, '注：以x轴为正向，逆时针方向为正，取值0~360（度）。'] + [None] * 8)
    path = "C:\\Users\\wuyan\\Desktop\\xiaosai\\A题_solution\\stage2\\result2.xlsx"
    wb.save(path)
    print(f"  result2.xlsx 已保存: {path}")


def write_result3(all_results):
    """result3.xlsx: 5机至多3弹对M1M2M3"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ['无人机编号', '无人机运动方向', '无人机运动速度 (m/s)', '烟幕干扰弹编号',
               '烟幕干扰弹投放点的x坐标 (m)', '烟幕干扰弹投放点的y坐标 (m)', '烟幕干扰弹投放点的z坐标 (m)',
               '烟幕干扰弹起爆点的x坐标 (m)', '烟幕干扰弹起爆点的y坐标 (m)', '烟幕干扰弹起爆点的z坐标 (m)',
               '有效干扰时长 (s)', '干扰的导弹编号']
    ws.append(headers)
    for uav_id in ['FY1', 'FY2', 'FY3', 'FY4', 'FY5']:
        if uav_id not in all_results:
            for j in range(3):
                ws.append([uav_id] + [None] * 11)
            continue
        missile_id, bomb_list = all_results[uav_id]
        th, vv, bombs, T, merged = bomb_list[0]
        for j, (tr, df) in enumerate(bombs):
            drop = uav_pos(uav_id, th, vv, tr)
            bp, _ = burst_point(uav_id, th, vv, tr, df)
            ws.append([uav_id, np.degrees(th) % 360, vv, j + 1,
                       drop[0], drop[1], drop[2], bp[0], bp[1], bp[2],
                       round(T, 4), missile_id])
        # 补空行至3弹
        for j in range(len(bombs), 3):
            ws.append([uav_id, None, None, j + 1] + [None] * 8)
    ws.append([None] * 12)
    ws.append([None, '注：以x轴为正向，逆时针方向为正，取值0~360（度）。'] + [None] * 10)
    path = "C:\\Users\\wuyan\\Desktop\\xiaosai\\A题_solution\\stage2\\result3.xlsx"
    wb.save(path)
    print(f"  result3.xlsx 已保存: {path}")


if __name__ == '__main__':
    # Q2
    print("=" * 60)
    print("Q2: 单弹优化")
    print("=" * 60)
    q2_params, q2_T, _, _ = solve_q2(n_particles=25, n_iter=30)

    # Q3
    print("\n" + "=" * 60)
    print("Q3: 三弹接力")
    print("=" * 60)
    q3_theta, q3_v, q3_bombs, q3_T, q3_merged = solve_q3(q2_params)
    write_result1(q3_theta, q3_v, q3_bombs, q3_T)

    # Q4
    print("\n" + "=" * 60)
    print("Q4: 三机协同")
    print("=" * 60)
    q4_results, q4_T, q4_merged = solve_q4(q2_params)
    write_result2(q4_results)

    # Q5
    print("\n" + "=" * 60)
    print("Q5: 多对多")
    print("=" * 60)
    q5_results, q5_T = solve_q5(q2_params, q4_results)
    write_result3(q5_results)

    # 汇总
    print("\n" + "=" * 60)
    print("汇总")
    print("=" * 60)
    print(f"Q1: 1.3920s (验证)")
    print(f"Q2: {q2_T:.4f}s")
    print(f"Q3: {q3_T:.4f}s")
    print(f"Q4: {q4_T:.4f}s")
    print(f"Q5: {q5_T:.4f}s (各弹独立求和)")
